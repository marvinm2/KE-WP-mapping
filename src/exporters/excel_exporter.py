"""
Excel export functionality with comprehensive data dictionary and multiple sheets
"""
import io
import logging
from datetime import datetime
from typing import Dict, List

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export dataset in Excel format with multiple sheets and data dictionary"""
    
    def __init__(self, database, metadata_manager):
        self.db = database
        self.metadata = metadata_manager
    
    def export(self, include_statistics: bool = True, include_metadata: bool = True) -> bytes:
        """Export dataset to Excel format with multiple sheets"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        conn = self.db.get_connection()
        try:
            # Get all mappings
            cursor = conn.execute("""
                SELECT id, ke_id, ke_title, wp_id, wp_title, connection_type, 
                       confidence_level, created_by, created_at, updated_at
                FROM mappings 
                ORDER BY created_at DESC
            """)
            mappings = [dict(row) for row in cursor.fetchall()]
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="307BBF", end_color="307BBF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sheet 1: Main Data
            ws_data = wb.create_sheet("Mappings Data")
            
            # Add headers
            headers = [
                "ID", "Key Event ID", "Key Event Title", "WikiPathway ID", 
                "WikiPathway Title", "Connection Type", "Confidence Level", 
                "Created By", "Created At", "Updated At"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data
            for row_idx, mapping in enumerate(mappings, 2):
                ws_data.cell(row=row_idx, column=1, value=mapping.get("id"))
                ws_data.cell(row=row_idx, column=2, value=mapping.get("ke_id"))
                ws_data.cell(row=row_idx, column=3, value=mapping.get("ke_title"))
                ws_data.cell(row=row_idx, column=4, value=mapping.get("wp_id"))
                ws_data.cell(row=row_idx, column=5, value=mapping.get("wp_title"))
                ws_data.cell(row=row_idx, column=6, value=mapping.get("connection_type"))
                ws_data.cell(row=row_idx, column=7, value=mapping.get("confidence_level"))
                ws_data.cell(row=row_idx, column=8, value=mapping.get("created_by"))
                ws_data.cell(row=row_idx, column=9, value=mapping.get("created_at"))
                ws_data.cell(row=row_idx, column=10, value=mapping.get("updated_at"))
                
                # Add borders to data cells
                for col in range(1, 11):
                    ws_data.cell(row=row_idx, column=col).border = border
            
            # Auto-adjust column widths
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Data Dictionary
            ws_dict = wb.create_sheet("Data Dictionary")
            
            # Data dictionary content
            dict_data = [
                ["Field Name", "Data Type", "Description", "Allowed Values", "Example"],
                ["id", "Integer", "Unique identifier for each mapping record", "Positive integers", "1, 2, 3..."],
                ["ke_id", "String", "Key Event identifier from AOP-Wiki database", "Format: 'KE XXXX'", "KE 1234"],
                ["ke_title", "String", "Full title/name of the Key Event", "Free text", "Oxidative stress in liver"],
                ["wp_id", "String", "WikiPathways pathway identifier", "Format: 'WPXXXX'", "WP1234"],
                ["wp_title", "String", "Full title/name of the WikiPathways pathway", "Free text", "Apoptosis signaling pathway"],
                ["connection_type", "String", "Type of biological relationship between KE and pathway", 
                 "causative, responsive, other, undefined", "causative"],
                ["confidence_level", "String", "Confidence in the mapping based on evidence strength", 
                 "low, medium, high", "high"],
                ["created_by", "String", "Username of person who created this mapping", "GitHub usernames", "researcher123"],
                ["created_at", "DateTime", "Timestamp when mapping was first created", "ISO 8601 format", "2025-01-15T10:30:00"],
                ["updated_at", "DateTime", "Timestamp when mapping was last modified", "ISO 8601 format", "2025-01-20T14:45:00"]
            ]
            
            for row_idx, row_data in enumerate(dict_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_dict.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    if row_idx == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Auto-adjust column widths for dictionary
            for column in ws_dict.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws_dict.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 3: Value Definitions
            ws_values = wb.create_sheet("Value Definitions")
            
            value_definitions = [
                ["Field", "Value", "Definition"],
                ["connection_type", "causative", "The pathway directly causes the Key Event to occur"],
                ["connection_type", "responsive", "The pathway responds to or is activated by the Key Event"],
                ["connection_type", "other", "Other defined biological relationship exists"],
                ["connection_type", "undefined", "Relationship type has not been determined"],
                ["", "", ""],
                ["confidence_level", "high", "Direct and specific biological link with strong experimental evidence"],
                ["confidence_level", "medium", "Partial or indirect biological relationship with moderate evidence"],
                ["confidence_level", "low", "Weak, speculative, or unclear biological connection"]
            ]
            
            for row_idx, row_data in enumerate(value_definitions, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_values.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    if row_idx == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws_values.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 80)
                ws_values.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 4: Statistics (if requested)
            if include_statistics:
                ws_stats = wb.create_sheet("Statistics")
                stats_data = self._generate_statistics_data(mappings)
                
                row = 1
                for category, data in stats_data.items():
                    # Category header
                    cell = ws_stats.cell(row=row, column=1, value=category.upper())
                    cell.font = Font(bold=True, size=14)
                    cell.fill = PatternFill(start_color="93D5F6", end_color="93D5F6", fill_type="solid")
                    row += 1
                    
                    if isinstance(data, dict):
                        for key, value in data.items():
                            ws_stats.cell(row=row, column=1, value=str(key))
                            ws_stats.cell(row=row, column=2, value=value)
                            row += 1
                    else:
                        ws_stats.cell(row=row, column=1, value=str(data))
                        row += 1
                    
                    row += 1  # Add spacing between categories
                
                # Auto-adjust column widths
                for column in ws_stats.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (TypeError, AttributeError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_stats.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 5: Metadata (if requested)
            if include_metadata:
                ws_meta = wb.create_sheet("Dataset Metadata")
                metadata = self.metadata.get_current_metadata()
                
                # Flatten metadata for display
                meta_items = [
                    ["Dataset Information", ""],
                    ["Title", metadata["titles"][0]["title"]],
                    ["Publisher", metadata["publisher"]],
                    ["Publication Year", str(metadata["publication_year"])],
                    ["Version", metadata["version"]],
                    ["Language", metadata["language"]],
                    ["License", metadata["rights_list"][0]["rights"] if metadata["rights_list"] else ""],
                    ["", ""],
                    ["Description", ""],
                    ["Abstract", next((d["description"] for d in metadata["descriptions"] 
                                     if d["description_type"] == "Abstract"), "")],
                    ["", ""],
                    ["Export Information", ""],
                    ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ["Record Count", str(len(mappings))],
                    ["Available Formats", ", ".join(metadata["formats"])]
                ]
                
                for row_idx, (key, value) in enumerate(meta_items, 1):
                    cell_key = ws_meta.cell(row=row_idx, column=1, value=key)
                    cell_value = ws_meta.cell(row=row_idx, column=2, value=value)
                    
                    if key and not value and key != "":  # Section headers
                        cell_key.font = Font(bold=True, size=12)
                        cell_key.fill = PatternFill(start_color="E6007E", end_color="E6007E", fill_type="solid")
                        cell_key.font = Font(bold=True, color="FFFFFF")
                
                # Auto-adjust column widths
                ws_meta.column_dimensions['A'].width = 20
                ws_meta.column_dimensions['B'].width = 80
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        finally:
            conn.close()
    
    def _generate_statistics_data(self, mappings: List[Dict]) -> Dict:
        """Generate statistical data for Excel sheet"""
        if not mappings:
            return {"Total Mappings": 0}
        
        # Basic statistics
        stats = {
            "Total Mappings": len(mappings)
        }
        
        # Confidence level distribution
        confidence_dist = {}
        for mapping in mappings:
            conf = mapping.get("confidence_level", "unknown")
            confidence_dist[conf] = confidence_dist.get(conf, 0) + 1
        
        stats["Confidence Level Distribution"] = confidence_dist
        
        # Connection type distribution
        connection_dist = {}
        for mapping in mappings:
            conn = mapping.get("connection_type", "unknown")
            connection_dist[conn] = connection_dist.get(conn, 0) + 1
        
        stats["Connection Type Distribution"] = connection_dist
        
        # Contributor statistics
        contributors = {}
        for mapping in mappings:
            contrib = mapping.get("created_by", "anonymous")
            contributors[contrib] = contributors.get(contrib, 0) + 1
        
        stats["Top Contributors"] = dict(list(sorted(contributors.items(), 
                                                   key=lambda x: x[1], reverse=True))[:10])
        
        # Temporal statistics
        years = {}
        for mapping in mappings:
            if mapping.get("created_at"):
                try:
                    year = mapping["created_at"][:4]
                    years[year] = years.get(year, 0) + 1
                except (TypeError, IndexError):
                    continue
        
        if years:
            stats["Yearly Distribution"] = years
        
        return stats